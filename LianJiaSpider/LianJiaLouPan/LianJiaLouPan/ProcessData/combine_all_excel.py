from openpyxl import Workbook
from openpyxl import load_workbook
import os

directory = 'Data'
wb = Workbook()
sheet = wb.active
# 添加第一行的标题
sheet.append(["城市", "楼盘名称", "物业类型", "楼盘价格", "楼盘位置", "经度", "纬度", "状态", "URL"])

file_cnt = 0
for dirpath, dirnames, filenames in os.walk(directory):     # 读取Data目录下所有的文件
    for filename in filenames:      # 遍历其中的所有文件
        file_cnt += 1
        path = dirpath + '\\' + filename
        print("正在读取文件：%s" % path)
        wb2 = load_workbook(path)
        ws2 = wb2.active
        row_cnt = 0
        cityname = filename[:filename.find('.')-6]
        for row in ws2.rows:
            row_cnt += 1
            if row_cnt == 1:        # 去掉每个Excel文件中的第一行
                continue
            ar = [cityname]
            for cell in row:
                ar.append(cell.value)       # 读取每一行的所有单元格
            sheet.append(ar)
wb.save('all.xlsx')     # 保存文件
